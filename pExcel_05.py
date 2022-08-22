from faker import Faker
import openpyxl
from openpyxl.chart import BarChart, Reference, LineChart


def init_test_data():
    """
    创建测试数据
    :return:
    """
    faker = Faker(locale='zh_CN')
    data = [
        ('月份', faker.name(), faker.name(),faker.name())
    ]

    for i in range(1, 13):
        data.append((
            f'{i}月', faker.random_number(digits=5), faker.random_number(digits=5), faker.random_number(digits=5)
        ))

    return data


def insert_data_to_sheet(ws):
    """
    将测试数据插入到工作表ws中
    :param ws: 工作表
    :return:
    """
    test_data = init_test_data()

    test_data = [

        ['Aliens', 2, 3, 4, 5, 6, 7],
        ['Humans', 10, 40, 50, 20, 10, 50],
    ]
    # 将测试数据写入到excel文件中
    for td in test_data:
        ws.append(td)


def create_bar_chart_excel(wb):
    """
    创建柱状图
    :param wb:
    :return:
    """
    ws = wb.create_sheet('柱状图')

    insert_data_to_sheet(ws)

    #数据选区
    data_refObj = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=13)
    labels_refObj = Reference(ws, min_col=1, min_row=2, max_row=13)

    bar_chart = BarChart()
    # bar_chart.type = 'bar' # 'col'
    bar_chart.style = 10  # 1-48
    bar_chart.title = '销售人员业绩表(2020)'
    bar_chart.x_axis.title = '月  份'
    bar_chart.y_axis.title = '销售额(万元)'
    bar_chart.width = 34
    bar_chart.height = 15

    # 填充数据
    bar_chart.add_data(data_refObj, titles_from_data=True)
    # 设置横轴显示内容
    bar_chart.set_categories(labels_refObj)

    # 将图形添加到工作表上
    ws.add_chart(bar_chart, 'F2')


def create_line_chart(wb):
    """
    创建折线图
    :param wb: 工作表
    :return:
    """
    ws = wb.create_sheet('拆线图')
    wb.active = ws

    insert_data_to_sheet(ws)

    line = LineChart()

    line.title = '销售人员业绩表(2020)'
    line.x_axis.title = '月  份'
    line.y_axis.title = '销售额(万元)'

    line.width = 34
    line.height = 15

    data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=13)
    labels = Reference(ws, min_col=1, min_row=2, max_row=13)

    line.add_data(data, titles_from_data=True)
    line.set_categories(labels)

    s1 = line.series[0]
    s1.graphicalProperties.line.solidFill = 'FA00AA'
    s1.graphicalProperties.line.dashStyle = 'sysDot'
    s1.smooth = True


    ws.add_chart(line, anchor='F2')


def handle_chart(ws):

    rows = [
        ['月份', '木材', '糖'],
        ['1月', 10000, 10],
        ['2月', 15002, 30],
        ['3月', 30000, 50],
        ['4月', 9500, 35],
        ['5月', 21000, 25],
        ['6月', 12352, 45],
        ['7月', 32009, 90],
    ]


    for row in rows:
        ws.append(row)

    c1 = BarChart()
    v1 = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=8)
    labels = Reference(ws, min_col=1, min_row=2, max_row=8)
    c1.add_data(v1, titles_from_data=True)
    c1.set_categories(labels)

    c1.x_axis.title = 'Days'
    c1.y_axis.title = 'Aliens'
    c1.y_axis.majorGridlines = None
    c1.title = 'Survey results'


    # Create a second chart
    c2 = LineChart()
    v2 = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=8)
    c2.add_data(v2, titles_from_data=True)
    c2.set_categories(labels)
    c2.y_axis.axId = 200
    c2.y_axis.title = "Humans"

    # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
    c1.y_axis.crosses = "max"
    c1 += c2

    ws.add_chart(c1, "D4")


if __name__ == '__main__':
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet()
    handle_chart(ws)
    # create_bar_chart_excel(wb)
    # create_line_chart(wb)

    wb.save('./pExcel_06.xlsx')