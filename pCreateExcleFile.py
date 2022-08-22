import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from faker import Faker


def init_student_infos():
    fake = Faker(locale='zh_CN')
    students = []
    for i in range(1000):
        stu = {'id': f'{i:04d}',
               'name': fake.name(),
               'chinese': fake.random_number(digits=2),
               'math': fake.random_number(digits=2),
               'english': fake.random_number(digits=2)
            }
        students.append(stu)

    return students


def set_title_row(sheet):
    """
    设置首行
    :param sheet:
    :return:
    """
    # 根据列的索引号，得到它对应的列的字母表示
    c = get_column_letter(7)
    print(c)
    # 合并单元格
    sheet.merge_cells(f'A1:{c}1')
    # 合并之后的单元格的坐标等于 合并之前的右上角的单元的坐标
    a1 = sheet['A1']
    # 设置单元格的值
    a1.value = '成绩表'

    font = Font(size=24, bold=True, color='0000FF')
    a1.font = font

    alignment = Alignment(horizontal='center', vertical='center')
    a1.alignment = alignment

    fill = PatternFill("solid", fgColor="DDDDDD")
    a1.fill = fill

    sheet.row_dimensions[1].height = 70


def set_header_row(sheet):
    """
    设置列头
    :param sheet:
    :return:
    """
    font = Font(size=18, italic=True)
    alignment = Alignment(horizontal='center', vertical='center')
    sheet['A2'] = '学号'
    sheet['B2'] = '姓名'
    sheet['C2'] = '语文'
    sheet['D2'] = '数学'
    sheet['E2'] = '英语'
    sheet['F2'] = '平均分'
    sheet['G2'] = '总分'
    for c in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
        sheet[f'{c}2'].font = font
        sheet[f'{c}2'].alignment = alignment

    sheet.row_dimensions[2].height = 50
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15


def set_data_cell(sheet):
    students = init_student_infos()
    for row in range(3, len(students) + 3):
        student = students[row - 3]
        sheet[f'A{row}'] = student['id']
        sheet[f'B{row}'] = student['name']
        sheet[f'C{row}'] = student['chinese']
        sheet[f'D{row}'] = student['math']
        sheet[f'E{row}'] = student['english']
        sheet[f'F{row}'] = f'=AVERAGE(C{row}:E{row})'
        sheet[f'G{row}'] = f'=SUM(C{row}:E{row})'
        # 设置数字显示1位小位
        sheet[f'F{row}'].number_format = '0.0' #numbers.FORMAT_NUMBER_00


def create_excel_file(file_path):
    # 获取Workbook对象
    wb = openpyxl.Workbook()
    # 得到活动工作表
    sheet = wb.active

    set_title_row(sheet)
    set_header_row(sheet)
    set_data_cell(sheet)

    # 冻结表格
    sheet.freeze_panes = 'B3'

    wb.save(file_path)


if __name__ == '__main__':
    create_excel_file('./成绩表.xlsx')