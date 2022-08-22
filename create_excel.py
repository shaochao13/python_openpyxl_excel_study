import openpyxl
import pathlib


def create_excel_file(file_path):
    # 创建一个默认的workbook对象，它会包含一个默认的工作表,名为"Sheet"
    wb = openpyxl.Workbook()
    # 获取所有的工作表名称
    print(wb.sheetnames)
    # 创建一个工作表，名称为Sheet_1形式
    wb.create_sheet()
    # 创建一个工作表，传入工作表的名称，工作表的位置
    wb.create_sheet(index=0, title='ABC')
    print(wb.sheetnames)
    # 保存workbook
    wb.save(file_path)


def remove_excel_sheet(file_path):
    # 先判断文件是否存在
    path = pathlib.PosixPath(file_path)
    if path.exists():
        # 根据文件路径，加载excel文件，得到一个workbook对象
        wb = openpyxl.load_workbook(path)
        #获取所有的工作表名称
        print(wb.sheetnames)
        # 删除指定工作表
        wb.remove_sheet(wb.get_sheet_by_name(wb.sheetnames[0]))
        # 保存workbook
        wb.save(path)


def update_excel_sheet(file_path):
    # 先判断文件是否存在
    path = pathlib.PosixPath(file_path)
    if path.exists():
        wb = openpyxl.load_workbook(file_path)
        # 获取所有的工作表的名称
        print(wb.sheetnames)
        # 获取活动的工作表
        sheet = wb.active
        # 或者根据名称获取要操作的工作表对象
        sheet = wb.get_sheet_by_name(wb.sheetnames[0])

        # 设置或者更新单元格的值
        sheet.cell(row=1, column=1).value = 'demo@demo.com'
        print(sheet['A1'].value)

        # 可以通过循环进行设置
        for r in range(1, 11):
            for c in range(1, 6):
                sheet.cell(row=r, column=c).value = r * c

        wb.save(path)
    else:
        print('文件不存在')


if __name__ == '__main__':
    # create_excel_file('/.abc.xlsx')
    remove_excel_sheet('./abc.xlsx')
    update_excel_sheet('./abc.xlsx')