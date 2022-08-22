from faker import Faker
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, numbers


def init_student_infos():
    """
    初始化测试数据
    :return:
    """
    faker = Faker(locale='zh_CN')
    students = []
    for i in range(1, 1001):
        stu = {
            'id': f'{i:04d}',
            'name': faker.name(),
            'chinese': faker.random_number(digits=2),
            'math': faker.random_number(digits=2),
            'english': faker.random_number(digits=2)
        }
        students.append(stu)
    return students


def set_first_row(sheet):
    """
    设置第一行
    :param sheet:
    :return:
    """
    c_char = get_column_letter(7)
    print(c_char)
    sheet.merge_cells(f'A1:{c_char}1')
    a1 = sheet['A1']
    a1.value = '成绩表'

    font = Font(size=24, bold=True, color='0000ff')
    a1.font = font

    alignment = Alignment(horizontal='center', vertical='center')
    a1.alignment = alignment

    fill = PatternFill('solid', fgColor='DDDDDD')
    a1.fill = fill

    sheet.row_dimensions[1].height = 70


def set_table_title_row(sheet):
    """
    设置第二行
    :param sheet:
    :return:
    """
    sheet['A2'] = '学号'
    sheet['B2'] = '姓名'
    sheet['C2'] = '语文'
    sheet['D2'] = '数学'
    sheet['E2'] = '英语'
    sheet['F2'] = '平均分'
    sheet['G2'] = '总分'

    font = Font(size=18, italic=True)
    alignment = Alignment(horizontal='center', vertical='center')

    for c in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
        sheet[f'{c}2'].font = font
        sheet[f'{c}2'].alignment = alignment

    # 设置行高
    sheet.row_dimensions[2].height = 50

    # 设置平均数、总分的列宽
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15


def set_data_cell(sheet):
    """
    设置数据
    :param sheet:
    :return:
    """
    students = init_student_infos()
    for row_number in range(3, len(students) + 3):
        stu = students[row_number - 3]
        sheet[f'A{row_number}'] = stu.get('id')
        sheet[f'B{row_number}'] = stu.get('name')
        sheet[f'C{row_number}'] = stu.get('chinese')
        sheet[f'D{row_number}'] = stu.get('math')
        sheet[f'E{row_number}'] = stu.get('english')
        sheet[f'F{row_number}'] = f'=AVERAGE(C{row_number}:E{row_number})'
        # 设置平均数只保留1位小数
        sheet[f'F{row_number}'].number_format = '0.0'  # numbers.FORMAT_NUMBER_00

        sheet[f'G{row_number}'] = f'=SUM(C{row_number}:E{row_number})'


def create_excel_file(file_path):
    """
    创建Excel文件
    :param file_path:
    :return:
    """
    wb = openpyxl.Workbook()
    sheet = wb.active

    set_first_row(sheet)
    set_table_title_row(sheet)

    set_data_cell(sheet)

    # 冻结第一列和前两行
    sheet.freeze_panes = 'B3'

    # 保存Excel
    wb.save(file_path)


if __name__ == '__main__':
    create_excel_file('./成绩表2.xlsx')
    # init_student_infos()