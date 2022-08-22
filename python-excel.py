import openpyxl

def read_excel_file():
    wb = openpyxl.load_workbook('./demo.xlsx')
    # print(type(wb))
    # sheets = wb.get_sheet_names()
    # print(sheets)
    sheet = wb.get_sheet_by_name('成绩表')
    # print(sheet)
    # cell = sheet['A1']
    # print(cell.value)
    # cell = sheet.cell(row=1, column=1)
    # print(cell.value, cell.row, cell.column, cell.coordinate)

    # #最大行号
    # print(sheet.max_row)
    # #最大列号
    # print(sheet.max_column)

    # for r in sheet.rows:
    #     print(r)

    # for c in sheet.columns:
    #     print(c)

    # cells = sheet['A2': 'E4']
    # for r in cells:
    #     for c in r:
    #         print(c.value, c.coordinate)



if __name__ == '__main__':
    read_excel_file()