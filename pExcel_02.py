import openpyxl
import pathlib
from openpyxl.styles import Font, Alignment


def create_excel_file(file_path):
    """
    创建Excel文件
    :param file_path: 文件路径
    :return:
    """
    # 创建一个默认的Workbook对象
    wb = openpyxl.Workbook()
    print(wb.sheetnames)
    # 创建新的工作表
    wb.create_sheet()
    print(wb.sheetnames)
    wb.create_sheet(title='Demo', index=0)
    print(wb.sheetnames)

    wb.save(file_path)

def remove_excel_sheet(file_path):
    """
    删除工作表
    :param file_path: Excel文件路径
    :return:
    """
    # 先判断文件是否存在
    path = pathlib.Path(file_path)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        print(wb.sheetnames)
        # 删除工作表
        wb.remove(wb[wb.sheetnames[2]])
        print(wb.sheetnames)
        wb.save(path)
    else:
        print('文件不存在')

def update_sheet_cell(file_path):
    """
    更新或者设置单元格的值
    :param file_path:
    :return:
    """
    # 先判断文件是否存在
    path = pathlib.Path(file_path)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        print(wb.sheetnames)
        # 获取到当前excel文件中的活动工作表
        sheet = wb.active
        print(sheet.title)
        # 通过名称拿到指定的工作表
        # sheet = wb[wb.sheetnames[0]]

        # sheet['A1'] = 'demo'
        # print(sheet.cell(row=1, column=1).value)

        # 10行 5列的表格
        for r in range(1, 11):
            for c in range(1, 6):
                sheet.cell(row=r, column=c).value = r * c

        wb.save(path)

    else:
        print('文件不存在')


def update_font_style(file_path):
    # 先判断文件是否存在
    path = pathlib.Path(file_path)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        font = Font(size=24, italic=True, color='FF0000')
        alignment = Alignment(horizontal='center', vertical='center')
        # sheet['A1'].font = font
        sheet['A3'].font = font
        sheet.merge_cells('A12:F14')
        sheet['A12'].value = 'DEMO'
        sheet['A12'].font = font
        sheet['A12'].alignment = alignment
        wb.save(path)

    else:
        print('文件不存在')

if __name__ == '__main__':
    # create_excel_file('./new_excel.xlsx')
    # remove_excel_sheet('./new_excel.xlsx')
    # update_sheet_cell('./new_excel.xlsx')
    update_font_style('./new_excel.xlsx')