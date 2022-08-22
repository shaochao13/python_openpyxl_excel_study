from faker import Faker
import openpyxl
from openpyxl.chart import BarChart, Reference, LineChart
from copy import deepcopy


def init_test_data():
    """
    创建测试数据
    :return:
    """
    faker = Faker(locale='zh_CN')
    data = [
        ('月份', faker.name(), faker.name(),faker.name())
    ]

    for i in range(1, 13):
        data.append((
            f'{i}月', faker.random_number(digits=5), faker.random_number(digits=5), faker.random_number(digits=5)
        ))

    return data


def insert_data_to_sheet(ws):
    test_data = init_test_data()
    # 将测试数据写入到excel文件中
    for td in test_data:
        ws.append(td)


def create_bar_chart_excel(wb):
    ws = wb.create_sheet('柱状图')

    insert_data_to_sheet(ws)

    #数据选区
    data_refObj = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=13)
    labels_refObj = Reference(ws, min_col=1, min_row=2, max_row=13)

    bar_chart = BarChart()
    # bar_chart.type = 'bar' # 'col'
    bar_chart.style = 10  # 1-48
    bar_chart.title = '销售人员业绩表(2020)'
    bar_chart.x_axis.title = '月  份'
    bar_chart.y_axis.title = '销售额(万元)'
    bar_chart.width = 34
    bar_chart.height = 15

    # 填充数据
    bar_chart.add_data(data_refObj, titles_from_data=True)
    # 设置横轴显示内容
    bar_chart.set_categories(labels_refObj)

    # chart3 = deepcopy(bar_chart)  # 深层复制 图表对象
    # chart3.title = 'Stacked Chart'  # 图表标题
    # chart3.style = 12  # 图表样式
    # chart3.type = 'bar'  # 注意：图表类型设置为‘col'来定义垂直条形图
    # chart3.grouping = 'stacked'  # 注意：图表分组属性设置为 二叠分堆放
    # chart3.overlap = 100  # 注意：将重叠overlap属性 设置为100 来定义使用堆叠图表。
    # ws.add_chart(chart3, 'A27')


    # 将图形添加到工作表上
    ws.add_chart(bar_chart, 'F2')


def create_line_chart(wb):
    ws = wb.create_sheet('折线图')
    wb.active = ws

    insert_data_to_sheet(ws)

    c1 = LineChart()  # 实例化创建 折线图表实例对象
    c1.width = 34
    c1.height = 15
    c1.smooth = True
    c1.title = '销售人员业绩表(2020)'
    c1.x_axis.title = '月  份'
    c1.y_axis.title = '销售额(万元)'
    data_refObj = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=13)
    labels_refObj = Reference(ws, min_col=1, min_row=2, max_row=13)
    c1.add_data(data_refObj, titles_from_data=True)  # 图表添加数据系列。
    c1.set_categories(labels_refObj)


    # 更改线型为三角线型
    # 使用：Chart.series 属性 来获取所有数据系列值。
    # s1 = c1.series[0]  # 获取折线图表的 数据系列
    # s1.marker.symbol = 'triangle'  # 更改线型为三角线型
    # s1.marker.graphicalProperties.solidFill = 'FF0000'  # 指定固体填充
    # s1.marker.graphicalProperties.line.solidFill = 'FF0000'  # 指定线的固体填充
    # s1.graphicalProperties.line.noFill = True  # 指定是否连线。
    # #
    # s2 = c1.series[1]
    # s2.graphicalProperties.line.solifFill = '00AAAA'
    # s2.graphicalProperties.line.dashStyle = 'sysDot'
    # s2.graphicalProperties.line.width = 100050
    # s2.smooth = True
    #
    # s3 = c1.series[2]
    # s3.smooth = True


    ws.add_chart(c1, 'F2')


if __name__ == '__main__':
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    create_bar_chart_excel(wb)
    create_line_chart(wb)
    wb.save('./pExcel_04.xlsx')