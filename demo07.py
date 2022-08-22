import openpyxl
from openpyxl.chart import StockChart, Reference, LineChart, BarChart
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.updown_bars import UpDownBars
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice


wb = openpyxl.load_workbook('/Users/kevin/Desktop/000001.xlsx')

ws_active = wb['000001']

ws_chart = wb.create_sheet('图形')
wb.active = ws_chart

s = StockChart()

data = Reference(ws_active, min_col=2, max_col=5, min_row=1, max_row=50)
s.add_data(data, titles_from_data=True)

labels = Reference(ws_active, min_col=1, min_row=2, max_row=50)
s.set_categories(labels)

for c in s.series:
    c.graphicalProperties.line.noFill = True

s.hiLowLines = ChartLines()
s.upDownBars = UpDownBars()

s.width = 35
s.height = 15


# 由于Excel中的BUG错误，仅当数据序列中的至少一个具有一些虚拟值时，才会显示高/低行
from openpyxl.chart.data_source import NumData, NumVal
pts = [NumVal(idx=i) for i in range(len(data) - 1)]
cache = NumData(pt=pts)
s.series[-1].val.numRef.numCache = cache

line = LineChart()
line.style = 10
data = Reference(ws_active, min_col=2, min_row=1, max_row=50)
line.add_data(data, titles_from_data=True)
line.title = '收盘价'

# Style the lines
s1 = line.series[0]
linePro = LineProperties(solidFill='FF0000', w=1)
s1.graphicalProperties.line = linePro

s += line

# bar = BarChart()
# data = Reference(ws_active, min_col=6, min_row=1, max_row=50)
# bar.add_data(data, titles_from_data=True)
# # 显示第二个y轴，如果不设置，将不会显示
# bar.y_axis.axId = 200
# # 设置网络线不显示
# bar.y_axis.majorGridlines = None
# bar.y_axis.crosses = 'max'
# bar.y_axis.title = '成交量'
# fill =  PatternFillProperties(prst="pct5")
# fill.foreground = ColorChoice(prstClr="red")
# fill.background = ColorChoice(prstClr="yellow")

# s2 = bar.series[0]
# s2.graphicalProperties.pattFill = fill

# s += bar

ws_chart.add_chart(s, 'B2')


wb.save('demo07.xlsx')